
from __future__ import annotations

import io
import numbers
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .quality import analyse_results
from .utils import autosize_width, format_br_number, format_brasilia_datetime, strip_numeric_noise_from_description

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
COLUMN_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
DARK_FONT = Font(color="000000", bold=True, name="Calibri", size=11)
NORMAL_FONT = Font(color="000000", bold=False, name="Calibri", size=11)
THIN = Side(style="thin", color="B7C9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BR_NUM_FMT = '#,##0.00'
BR_INT_FMT = '0'
TABLE_COL_WIDTHS = {"A": 8.43, "B": 16.0, "C": 36.0, "D": 20.0, "E": 13.0, "F": 13.0}


NUMERIC_HEADER_KEYWORDS = (
    "saídas", "saidas", "baixas", "depreciação", "depreciacao",
    "acumulada", "valor", "vlr", "líq", "liq", "contábil", "contabil",
    "saldo", "entradas", "reavaliação", "reavaliacao", "total",
    "imb010", "imb025", "imb037", "diferença", "diferenca", "quantidade", "qtd",
    "doação", "doacao"
)
NON_MONETARY_HEADERS = {"grupo", "conta contábil", "conta contabil"}


def _is_numeric_value(value) -> bool:
    return isinstance(value, numbers.Number) and not isinstance(value, bool)


def _is_numeric_header(header: object) -> bool:
    if header is None:
        return False
    text = str(header).strip().lower()
    if not text or text in NON_MONETARY_HEADERS:
        return False
    return any(keyword in text for keyword in NUMERIC_HEADER_KEYWORDS)


def _apply_numeric_display_format(ws, header_row: int, data_start_row: int | None = None, numeric_columns: set[str] | None = None) -> None:
    """Aplica formato numérico brasileiro às colunas monetárias/quantitativas.

    A formatação é somente visual: os valores permanecem como números no Excel,
    preservando fórmulas, somatórios, filtros e totalizações.
    """
    numeric_columns = numeric_columns or set()
    data_start_row = data_start_row or header_row + 1
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(header_row, col_idx).value
        header_text = str(header_value or "").strip()
        should_format = header_text in numeric_columns or _is_numeric_header(header_text)
        if not should_format:
            continue
        for row_idx in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if _is_numeric_value(cell.value):
                cell.number_format = BR_NUM_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=cell.alignment.wrap_text)


REPORT_COLUMNS = [
    "Grupo",
    "Conta Contábil",
    "Descrição do Grupo",
    "Saídas/Grupo",
    "Dep. Acumulada",
    "Vlr. Liq. Contábil",
]
MEMORY_COLUMNS = REPORT_COLUMNS + ["Fonte Saídas", "Fonte Depreciação", "Competência"]


def load_pcasp_map(path: str | Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    normalized = {str(col).strip().lower(): col for col in df.columns}
    grupo_col = next((v for k, v in normalized.items() if "grupo" in k), None)
    conta_col = next((v for k, v in normalized.items() if "conta" in k), None)
    if not grupo_col or not conta_col:
        raise ValueError("A base de Conta Contábil deve conter colunas de grupo e conta.")
    out = df[[grupo_col, conta_col]].copy()
    out.columns = ["Grupo", "Conta Contábil"]
    out["Grupo"] = pd.to_numeric(out["Grupo"], errors="coerce")
    out = out.dropna(subset=["Grupo"])
    out["Grupo"] = out["Grupo"].astype(int)
    out["Conta Contábil"] = out["Conta Contábil"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    out = out.groupby("Grupo", as_index=False)["Conta Contábil"].agg(
        lambda vals: " / ".join(dict.fromkeys(v for v in vals if v and v.lower() != "nan"))
    )
    return out


def consolidate_by_group(df: pd.DataFrame, sum_columns: Iterable[str], first_columns: Iterable[str]) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    agg = {col: "sum" for col in sum_columns}
    agg.update({col: "first" for col in first_columns})
    out = df.groupby("Grupo", as_index=False).agg(agg)
    return out.sort_values("Grupo").reset_index(drop=True)


def build_outputs(
    rsp_df: pd.DataFrame,
    dep_df: pd.DataFrame,
    pcasp_df: pd.DataFrame,
    rsp_meta: dict,
    dep_meta: dict,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    logs: list[dict[str, str]] = []

    # A coluna Saídas/Grupo representa, a partir da v10, somente as demais saídas,
    # excluídas as transferências. As colunas auxiliares mantêm rastreabilidade
    # entre o total original do SICPAT e a base efetivamente considerada.
    rsp_sum_cols = ["Saídas/Grupo"]
    for optional_col in ["Doação", "Saídas por Transferência", "Total Saídas Original"]:
        if optional_col in rsp_df.columns:
            rsp_sum_cols.append(optional_col)
    rsp_work = consolidate_by_group(
        rsp_df,
        sum_columns=rsp_sum_cols,
        first_columns=[
            "Descrição do Grupo",
            "Saldo Anterior",
            "Entradas",
            "Reavaliação",
            "Red.Vlr.Rec.",
            "Saldo Atual",
            "Depreciação do RSP",
            "Vlr. Liq. Cont. do RSP",
        ],
    )
    dep_work = consolidate_by_group(
        dep_df,
        sum_columns=["Dep. Acumulada (IMB010)"],
        first_columns=["Descrição do Grupo"],
    )

    if len(rsp_work) != len(rsp_df):
        logs.append({"Tipo": "Ajuste", "Mensagem": "Registros repetidos do Relatório Sintético Patrimonial foram consolidados por grupo."})
    if len(dep_work) != len(dep_df):
        logs.append({"Tipo": "Ajuste", "Mensagem": "Registros repetidos da Depreciação Acumulada foram consolidados por grupo."})
    if "Saídas por Transferência" in rsp_work.columns:
        transf_total = float(pd.to_numeric(rsp_work["Saídas por Transferência"], errors="coerce").fillna(0).sum())
        if abs(transf_total) > 0.01:
            logs.append({"Tipo": "Regra de apuração", "Mensagem": f"Saídas por transferência desconsideradas na base de cálculo: R$ {format_br_number(transf_total)}."})
    if "Doação" not in rsp_work.columns:
        rsp_work["Doação"] = 0.0
    rsp_work["Doação"] = pd.to_numeric(rsp_work["Doação"], errors="coerce").fillna(0.0).clip(lower=0)
    doacao_total = float(rsp_work["Doação"].sum())
    if abs(doacao_total) > 0.01:
        logs.append({"Tipo": "Regra de apuração", "Mensagem": f"Saídas por doação segregadas na aba Registro SIAFI Web: R$ {format_br_number(doacao_total)}."})

    merged = rsp_work.merge(dep_work[["Grupo", "Dep. Acumulada (IMB010)"]], on="Grupo", how="left")
    merged = merged.merge(pcasp_df, on="Grupo", how="left")
    merged["Conta Contábil"] = merged["Conta Contábil"].fillna("")
    merged["Dep. Acumulada (IMB010)"] = merged["Dep. Acumulada (IMB010)"].fillna(0.0)
    merged["Vlr. Liq. Contábil (IMB025)"] = merged["Saídas/Grupo"] - merged["Dep. Acumulada (IMB010)"]

    final_df = merged.loc[
        merged["Saídas/Grupo"] != 0,
        ["Grupo", "Conta Contábil", "Descrição do Grupo", "Saídas/Grupo", "Dep. Acumulada (IMB010)", "Vlr. Liq. Contábil (IMB025)"],
    ].copy()
    final_df = consolidate_by_group(
        final_df,
        sum_columns=["Saídas/Grupo", "Dep. Acumulada (IMB010)", "Vlr. Liq. Contábil (IMB025)"],
        first_columns=["Conta Contábil", "Descrição do Grupo"],
    )
    final_df = final_df.rename(
        columns={
            "Dep. Acumulada (IMB010)": "Dep. Acumulada",
            "Vlr. Liq. Contábil (IMB025)": "Vlr. Liq. Contábil",
        }
    )
    final_df = final_df[REPORT_COLUMNS]
    final_df["Descrição do Grupo"] = final_df["Descrição do Grupo"].fillna("").astype(str).map(strip_numeric_noise_from_description).str.upper()

    doacao_por_grupo = merged.set_index("Grupo")["Doação"].to_dict() if "Doação" in merged.columns else {}
    saidas_por_grupo = merged.set_index("Grupo")["Saídas/Grupo"].to_dict() if "Saídas/Grupo" in merged.columns else {}
    final_df.attrs["doacao_por_grupo"] = doacao_por_grupo
    final_df.attrs["saidas_por_grupo"] = saidas_por_grupo
    grupos_inconsistentes = []
    for grupo, doacao in doacao_por_grupo.items():
        saidas = float(saidas_por_grupo.get(grupo, 0.0) or 0.0)
        if float(doacao or 0.0) - saidas > 0.01:
            grupos_inconsistentes.append(int(grupo))
    if grupos_inconsistentes:
        logs.append({
            "Tipo": "Atenção",
            "Mensagem": "Doação superior ao total de saídas em grupo(s) "
            f"{grupos_inconsistentes}. A situação IMB025 foi limitada a zero para evitar valor negativo."
        })

    missing = final_df.loc[final_df["Conta Contábil"].eq(""), "Grupo"].tolist()
    if missing:
        logs.append({"Tipo": "Atenção", "Mensagem": f"Grupos sem Conta Contábil mapeada: {missing}"})
    if final_df.empty:
        logs.append({"Tipo": "Atenção", "Mensagem": "Nenhum grupo com saídas diferente de zero foi localizado."})
    if rsp_meta.get("competence") and dep_meta.get("competence") and rsp_meta.get("competence") != dep_meta.get("competence"):
        logs.append(
            {
                "Tipo": "Atenção",
                "Mensagem": f"Competências divergentes detectadas: RSP={rsp_meta.get('competence')} e Depreciação={dep_meta.get('competence')}",
            }
        )

    memoria = merged[["Grupo", "Conta Contábil", "Descrição do Grupo", "Saídas/Grupo", "Dep. Acumulada (IMB010)", "Vlr. Liq. Contábil (IMB025)"]].copy()
    memoria = memoria.rename(
        columns={
            "Dep. Acumulada (IMB010)": "Dep. Acumulada",
            "Vlr. Liq. Contábil (IMB025)": "Vlr. Liq. Contábil",
        }
    )
    memoria["Descrição do Grupo"] = memoria["Descrição do Grupo"].fillna("").astype(str).map(strip_numeric_noise_from_description).str.upper()
    memoria["Fonte Saídas"] = rsp_meta.get("filename", "")
    memoria["Fonte Depreciação"] = dep_meta.get("filename", "")
    memoria["Competência"] = rsp_meta.get("competence") or dep_meta.get("competence") or ""
    memoria = memoria[MEMORY_COLUMNS]

    logs_df = pd.DataFrame(logs or [{"Tipo": "OK", "Mensagem": "Processamento concluído sem inconsistências críticas."}])
    meta_df = pd.DataFrame(
        [
            {"Campo": "UG", "Valor": rsp_meta.get("ug_code") or dep_meta.get("ug_code") or ""},
            {"Campo": "Unidade Gestora", "Valor": rsp_meta.get("ug_name") or dep_meta.get("ug_name") or ""},
            {"Campo": "Período", "Valor": rsp_meta.get("period_label") or ""},
            {"Campo": "Competência", "Valor": rsp_meta.get("competence") or dep_meta.get("competence") or ""},
            {"Campo": "Arquivo RSP", "Valor": rsp_meta.get("filename") or ""},
            {"Campo": "Arquivo Depreciação", "Valor": dep_meta.get("filename") or ""},
            {"Campo": "Regra de apuração", "Valor": "Desconsidera saídas por transferência; utiliza a coluna SAÍDAS do quadro de saídas do SICPAT."},
            {"Campo": "Gerado em", "Valor": format_brasilia_datetime()},
        ]
    )
    return final_df, memoria, logs_df, meta_df


def _final_with_total(final_df: pd.DataFrame) -> pd.DataFrame:
    if final_df.empty:
        return final_df.copy()
    total = pd.DataFrame(
        [
            {
                "Grupo": "Total",
                "Conta Contábil": "",
                "Descrição do Grupo": "",
                "Saídas/Grupo": final_df["Saídas/Grupo"].sum(),
                "Dep. Acumulada": final_df["Dep. Acumulada"].sum(),
                "Vlr. Liq. Contábil": final_df["Vlr. Liq. Contábil"].sum(),
            }
        ]
    )
    return pd.concat([final_df, total], ignore_index=True)


def build_siafi_web_df(final_df: pd.DataFrame) -> pd.DataFrame:
    """Monta a aba Registro SIAFI Web.

    A aba utiliza a mesma base de grupos do Relatório Sintético, mas segrega as
    saídas por doação em IMB037. O Relatório Sintético permanece sem alteração.
    """
    doacao_por_grupo = final_df.attrs.get("doacao_por_grupo", {}) or {}
    saidas_por_grupo = final_df.attrs.get("saidas_por_grupo", {}) or {}

    registros: list[dict] = []
    for _, row in final_df.iterrows():
        grupo = row.get("Grupo")
        try:
            grupo_key = int(float(grupo))
        except Exception:
            grupo_key = grupo

        saidas = float(saidas_por_grupo.get(grupo_key, row.get("Saídas/Grupo", 0.0)) or 0.0)
        doacao = max(float(doacao_por_grupo.get(grupo_key, 0.0) or 0.0), 0.0)
        imb025 = max(saidas - doacao, 0.0)

        registros.append(
            {
                "Conta Contábil": row.get("Conta Contábil", ""),
                "Situação IMB010": max(float(row.get("Dep. Acumulada", 0.0) or 0.0), 0.0),
                "Situação IMB025": imb025,
                "Situação IMB037": doacao,
            }
        )

    base = pd.DataFrame(registros, columns=["Conta Contábil", "Situação IMB010", "Situação IMB025", "Situação IMB037"])
    if base.empty:
        return base
    total = pd.DataFrame(
        [
            {
                "Conta Contábil": "Total",
                "Situação IMB010": base["Situação IMB010"].sum(),
                "Situação IMB025": base["Situação IMB025"].sum(),
                "Situação IMB037": base["Situação IMB037"].sum(),
            }
        ]
    )
    return pd.concat([base, total], ignore_index=True)


def _style_cells(ws, row: int, fill, font, center: bool = False, height: float | None = None) -> None:
    if height is not None:
        ws.row_dimensions[row].height = height
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)


def _apply_reference_widths(ws) -> None:
    for col, width in TABLE_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width


def _style_final_sheet(ws) -> None:
    max_col, last_row = ws.max_column, ws.max_row
    # Header rows
    _style_cells(ws, 1, HEADER_FILL, WHITE_FONT, center=False, height=19)
    _style_cells(ws, 2, SUBHEADER_FILL, DARK_FONT, center=False, height=18)
    _style_cells(ws, 3, SUBHEADER_FILL, DARK_FONT, center=False, height=18)
    _style_cells(ws, 4, COLUMN_FILL, DARK_FONT, center=True, height=18)

    numeric_cols = {4, 5, 6}
    for row_idx in range(5, last_row + 1):
        ws.row_dimensions[row_idx].height = 15
        is_total = row_idx == last_row
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.font = DARK_FONT if is_total else NORMAL_FONT
            cell.fill = TOTAL_FILL if is_total else PatternFill(fill_type=None)
            cell.alignment = Alignment(
                horizontal="center" if col_idx == 1 else ("right" if col_idx in numeric_cols else "left"),
                vertical="center",
                wrap_text=(col_idx == 3),
            )
            if col_idx in numeric_cols and isinstance(cell.value, (int, float)):
                cell.number_format = BR_NUM_FMT
        if is_total:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            merged = ws.cell(row=row_idx, column=1)
            merged.value = "Total"
            merged.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A4:{get_column_letter(max_col)}{last_row}"
    _apply_numeric_display_format(ws, header_row=4, data_start_row=5, numeric_columns={"Saídas/Grupo", "Dep. Acumulada", "Vlr. Liq. Contábil"})
    _apply_reference_widths(ws)


def _style_siafi_sheet(ws) -> None:
    max_col, last_row = ws.max_column, ws.max_row
    _style_cells(ws, 1, HEADER_FILL, WHITE_FONT, height=19)
    _style_cells(ws, 2, SUBHEADER_FILL, DARK_FONT, height=18)
    _style_cells(ws, 3, COLUMN_FILL, DARK_FONT, center=True, height=18)
    for row_idx in range(4, last_row + 1):
        ws.row_dimensions[row_idx].height = 15
        is_total = row_idx == last_row
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.font = DARK_FONT if is_total else NORMAL_FONT
            cell.fill = TOTAL_FILL if is_total else PatternFill(fill_type=None)
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")
            if col_idx >= 2 and isinstance(cell.value, (int, float)):
                cell.number_format = BR_NUM_FMT
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A3:{get_column_letter(max_col)}{last_row}"
    _apply_numeric_display_format(ws, header_row=3, data_start_row=4, numeric_columns={"Situação IMB010", "Situação IMB025", "Situação IMB037"})
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = autosize_width(
            [c.value for c in col], minimum=16, maximum=24
        )


def _style_aux_sheet(
    ws,
    numeric_columns: set[str] | None = None,
    right_align_body_indexes: set[int] | None = None,
) -> None:
    numeric_columns = numeric_columns or set()
    right_align_body_indexes = set(right_align_body_indexes or set())
    max_col = ws.max_column
    _style_cells(ws, 1, COLUMN_FILL, DARK_FONT, center=True, height=18)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 15
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            horizontal = "right" if col_idx in right_align_body_indexes else "left"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
            if ws.cell(1, col_idx).value in numeric_columns and isinstance(cell.value, (int, float)):
                cell.number_format = BR_NUM_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{ws.max_row}"
    _apply_numeric_display_format(ws, header_row=1, data_start_row=2, numeric_columns=numeric_columns)
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = autosize_width(
            [c.value for c in col], minimum=14, maximum=28
        )


def generate_excel(
    final_df: pd.DataFrame,
    memoria_df: pd.DataFrame,
    logs_df: pd.DataFrame,
    meta_df: pd.DataFrame,
    header: dict,
) -> bytes:
    final_export = _final_with_total(final_df.copy())
    siafi = build_siafi_web_df(final_df)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        final_export.to_excel(writer, sheet_name="Relatório Sintético", index=False, startrow=3)
        siafi.to_excel(writer, sheet_name="Registro SIAFI Web", index=False, startrow=2)
        memoria_df.to_excel(writer, sheet_name="Memória de Cálculo", index=False)
        logs_df.to_excel(writer, sheet_name="Logs", index=False)
        meta_df.to_excel(writer, sheet_name="Metadados", index=False)
        analyse_results(final_df, memoria_df).to_excel(writer, sheet_name="Análise de Consistência", index=False)

        wb = writer.book
        ws = wb["Relatório Sintético"]
        ws["A1"] = "Relatório Auxiliar para Registro de Baixas de Bens Móveis"
        ws["A2"] = f"UG: {header.get('ug_code', '')} - {header.get('ug_name', '')}"
        ws["A3"] = f"Competência: {header.get('competence', '')} | Gerado em: {format_brasilia_datetime()}"
        for row in (1, 2, 3):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)
        _style_final_sheet(ws)

        ws2 = wb["Registro SIAFI Web"]
        ws2["A1"] = "Registro SIAFI Web"
        ws2["A2"] = f"Competência: {header.get('competence', '')} | Gerado em: {format_brasilia_datetime()}"
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws2.max_column)
        ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws2.max_column)
        _style_siafi_sheet(ws2)

        _style_aux_sheet(wb["Memória de Cálculo"], {"Saídas/Grupo", "Dep. Acumulada", "Vlr. Liq. Contábil"})
        _style_aux_sheet(wb["Logs"], right_align_body_indexes={2, 3})
        _style_aux_sheet(wb["Metadados"], right_align_body_indexes={2, 3})
        _style_aux_sheet(wb["Análise de Consistência"])
        wb._sheets = [wb["Relatório Sintético"], wb["Registro SIAFI Web"], wb["Memória de Cálculo"], wb["Análise de Consistência"], wb["Logs"], wb["Metadados"]]
    output.seek(0)
    return output.getvalue()


def _pdf_header_footer(canvas: Canvas, doc, header: dict) -> None:
    canvas.saveState()
    page_w, page_h = landscape(A4)
    left = doc.leftMargin
    right = page_w - doc.rightMargin
    top = page_h - 0.75 * cm

    header_lines = [
        ("UNIVERSIDADE FEDERAL DE MINAS GERAIS", 11, "Helvetica-Bold"),
        ("SISTEMA AUXILIAR DE DESFAZIMENTO PATRIMONIAL - SADPat", 9.5, "Helvetica-Bold"),
        ((header.get("ug_name") or "HOSPITAL DAS CLÍNICAS").upper(), 9.5, "Helvetica-Bold"),
        ("RELATÓRIO AUXILIAR PARA REGISTRO CONTÁBIL DE BAIXAS DE BENS MÓVEIS", 9.5, "Helvetica-Bold"),
    ]

    y = top
    for text, size, font in header_lines:
        canvas.setFont(font, size)
        canvas.drawCentredString((left + right) / 2, y, text)
        y -= 0.38 * cm

    canvas.setFont("Helvetica", 7.6)
    canvas.drawString(left, top - 0.12 * cm, f"UG: {header.get('ug_code', '')}")
    canvas.drawString(left, top - 0.50 * cm, f"Competência: {(header.get('competence') or '').upper()}")
    canvas.drawRightString(right, top - 0.12 * cm, f"Gerado em: {format_brasilia_datetime()}")
    canvas.drawRightString(right, top - 0.50 * cm, f"Página: {canvas.getPageNumber()}")
    canvas.line(left, top - 1.66 * cm, right, top - 1.66 * cm)

    canvas.setFont("Helvetica", 7)
    canvas.drawRightString(right, 0.65 * cm, "Fonte: SICPAT - SISTEMA DE CONTROLE PATRIMONIAL")
    canvas.restoreState()


def _pdf_paragraph(value, style: ParagraphStyle) -> Paragraph:
    if pd.isna(value):
        value = ""
    return Paragraph(str(value), style)


def _format_pdf_value(value, header_name: str) -> str:
    if pd.isna(value):
        return ""
    normalized_header = str(header_name).strip().lower()
    if normalized_header == "grupo":
        if str(value).upper() == "TOTAL":
            return "TOTAL"
        try:
            return str(int(float(value)))
        except Exception:
            return str(value)
    if _is_numeric_header(header_name):
        try:
            return format_br_number(float(value), 2)
        except Exception:
            return str(value)
    return str(value)


def _pdf_section_title(title: str, style: ParagraphStyle) -> list:
    return [Paragraph(title, style), Spacer(1, 0.10 * cm)]


def _generic_pdf_table(
    df: pd.DataFrame,
    cell: ParagraphStyle,
    cell_center: ParagraphStyle,
    cell_right: ParagraphStyle,
    col_widths: list | None = None,
    total_label_columns: int = 1,
    body_right_align_columns: set[str] | None = None,
    body_right_align_indexes: set[int] | None = None,
) -> Table:
    cols = list(df.columns)
    body_right_align_columns = {str(c).strip().lower() for c in (body_right_align_columns or set())}
    # Índices em base 1 para facilitar regras visuais solicitadas pelo usuário
    # sem depender do nome técnico das colunas. O cabeçalho permanece centralizado.
    body_right_align_indexes = set(body_right_align_indexes or set())
    data = [[Paragraph(str(c), cell_center) for c in cols]]

    if df.empty:
        data.append([Paragraph("Sem dados para exibição", cell_center)] + [Paragraph("", cell_center) for _ in cols[1:]])
    else:
        for idx, row in df.iterrows():
            is_total = idx == len(df) - 1 and str(row.iloc[0]).strip().upper() in {"TOTAL", "Total".upper()}
            row_cells = []
            for col_idx, col in enumerate(cols, start=1):
                normalized_col = str(col).strip().lower()
                style = cell_right if (
                    _is_numeric_header(col)
                    or normalized_col in body_right_align_columns
                    or col_idx in body_right_align_indexes
                ) else cell
                if normalized_col in {"grupo", "conta contábil", "conta contabil"}:
                    style = cell_center
                row_cells.append(Paragraph(_format_pdf_value(row[col], col), style))
            data.append(row_cells)

    if col_widths is None:
        usable_width = landscape(A4)[0] - 2.0 * cm
        col_widths = [usable_width / max(len(cols), 1)] * len(cols)

    table = Table(data, repeatRows=1, colWidths=col_widths, splitByRow=1)
    style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.0),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B7C9D6")),
            ("BOX", (0, 0), (-1, -1), 0.55, colors.HexColor("#B7C9D6")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3.5),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7FBFE")]),
        ]
    )

    if not df.empty:
        last_idx = len(data) - 1
        first_cell = str(df.iloc[-1, 0]).strip().upper()
        if first_cell == "TOTAL":
            if total_label_columns > 1 and len(cols) >= total_label_columns:
                style.add("SPAN", (0, last_idx), (total_label_columns - 1, last_idx))
                style.add("ALIGN", (0, last_idx), (total_label_columns - 1, last_idx), "CENTER")
            style.add("BACKGROUND", (0, last_idx), (-1, last_idx), colors.HexColor("#E2F0D9"))
            style.add("FONTNAME", (0, last_idx), (-1, last_idx), "Helvetica-Bold")

    table.setStyle(style)
    return table


def _logs_metadata_pdf_df(logs_df: pd.DataFrame, meta_df: pd.DataFrame) -> pd.DataFrame:
    """Consolida Logs e Metadados em quadro único de três colunas para o PDF.

    A estrutura visual fica padronizada como Quadro/Campo/Valor. Dessa forma, as
    duas colunas de dados solicitadas pelo usuário correspondem à segunda e à
    terceira colunas do quadro, permitindo alinhamento à direita no corpo da
    tabela sem alterar o conteúdo informacional de logs e metadados.
    """
    frames: list[pd.DataFrame] = []

    if logs_df is not None and not logs_df.empty:
        logs = logs_df.copy().astype(str)
        campo = logs["Tipo"] if "Tipo" in logs.columns else pd.Series(["Log"] * len(logs), index=logs.index)
        valor = logs["Mensagem"] if "Mensagem" in logs.columns else logs.astype(str).agg(" | ".join, axis=1)
        frames.append(pd.DataFrame({"Quadro": "Logs", "Campo": campo, "Valor": valor}))

    if meta_df is not None and not meta_df.empty:
        meta = meta_df.copy().astype(str)
        if {"Campo", "Valor"}.issubset(meta.columns):
            frames.append(pd.DataFrame({"Quadro": "Metadados", "Campo": meta["Campo"], "Valor": meta["Valor"]}))
        else:
            frames.append(pd.DataFrame({
                "Quadro": "Metadados",
                "Campo": meta.columns.astype(str),
                "Valor": meta.astype(str).agg(" | ".join, axis=0).values,
            }))

    if not frames:
        return pd.DataFrame({"Quadro": ["Logs e Metadados"], "Campo": ["Informação"], "Valor": ["Sem dados para exibição"]})

    return pd.concat(frames, ignore_index=True).fillna("").astype(str)


def generate_pdf(
    final_df: pd.DataFrame,
    header: dict,
    logs_df: pd.DataFrame | None = None,
    meta_df: pd.DataFrame | None = None,
) -> bytes:
    """Gera o Relatório de Baixas em PDF.

    O PDF consolida, em páginas próprias, os quadros:
    1. Relatório Sintético;
    2. Registro SIAFI Web;
    3. Logs e Metadados.

    A geração reutiliza os DataFrames já apurados, sem alterar as regras de
    cálculo, extração, validação ou exportação para Excel.
    """
    pdf_df = _final_with_total(final_df.copy())
    siafi_df = build_siafi_web_df(final_df)
    logs_meta_df = _logs_metadata_pdf_df(logs_df, meta_df)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.0 * cm,
        rightMargin=1.0 * cm,
        topMargin=4.1 * cm,
        bottomMargin=1.3 * cm,
        title="Relatório de Baixas",
    )

    styles = getSampleStyleSheet()
    section_title = ParagraphStyle(
        "section_title",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    cell = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.7,
        leading=8.2,
        alignment=TA_LEFT,
    )
    cell_center = ParagraphStyle("cell_center", parent=cell, alignment=TA_CENTER)
    cell_right = ParagraphStyle("cell_right", parent=cell, alignment=TA_RIGHT)

    story = []

    # Página 1 — Relatório Sintético
    story.extend(_pdf_section_title("QUADRO 1 — RELATÓRIO SINTÉTICO", section_title))
    final_col_widths = [1.65 * cm, 3.2 * cm, 9.7 * cm, 3.9 * cm, 3.2 * cm, 3.6 * cm]
    story.append(
        _generic_pdf_table(
            pdf_df,
            cell,
            cell_center,
            cell_right,
            col_widths=final_col_widths,
            total_label_columns=3,
        )
    )

    # Página 2 — Registro SIAFI Web
    story.append(PageBreak())
    story.extend(_pdf_section_title("QUADRO 2 — REGISTRO SIAFI WEB", section_title))
    siafi_col_widths = [6.0 * cm, 5.0 * cm, 5.0 * cm, 5.0 * cm]
    story.append(
        _generic_pdf_table(
            siafi_df,
            cell,
            cell_center,
            cell_right,
            col_widths=siafi_col_widths,
            total_label_columns=1,
        )
    )

    # Página 3 — Logs e Metadados
    story.append(PageBreak())
    story.extend(_pdf_section_title("QUADRO 3 — LOGS E METADADOS", section_title))
    usable_width = landscape(A4)[0] - 2.0 * cm
    logs_col_widths = [3.1 * cm, 5.0 * cm, usable_width - 8.1 * cm]
    story.append(
        _generic_pdf_table(
            logs_meta_df,
            cell,
            cell_center,
            cell_right,
            col_widths=logs_col_widths,
            total_label_columns=1,
            body_right_align_columns={"Campo", "Valor"},
            body_right_align_indexes={2, 3},
        )
    )

    doc.build(
        story,
        onFirstPage=lambda c, d: _pdf_header_footer(c, d, header),
        onLaterPages=lambda c, d: _pdf_header_footer(c, d, header),
    )
    buffer.seek(0)
    return buffer.getvalue()


